import pandas as pd#pandasのインポート
import datetime#元データの日付処理のためにインポート
import numpy as np
from sklearn import ensemble, model_selection
from sklearn.model_selection import train_test_split, GridSearchCV#データ分割用
from sklearn.model_selection import RandomizedSearchCV#ランダムサーチ
from sklearn.model_selection import cross_val_score # クロスバリデーション用
from sklearn.svm import SVC # SVM用
from sklearn import metrics # 精度検証用
import pandas as pd
from sklearn.metrics import classification_report, accuracy_score, confusion_matrix,precision_score,recall_score,f1_score
from sklearn.metrics import r2_score#決定係数求める用
import matplotlib.pyplot as plt
from sklearn.model_selection import train_test_split
import scipy.stats
import glob
import os
import openpyxl

for l in range(3):
    if l == 0:
        files = glob.glob('C:\\Users\\rikua\\Documents\\all_csv_mini_five\\*')
    elif l == 1:
        files = glob.glob('C:\\Users\\rikua\\Documents\\all_csv_mini_five\\*')
    else :
        files = glob.glob('C:\\Users\\rikua\\Documents\\all_csv_mini_five\\*')
        
    df_list_test = []
    df_list_train = []
    files = glob.glob('C:\\Users\\rikua\\Documents\\30s_all\\*')
    #test = random.sample(files,4)
    test = [files[0],files[5],files[11],files[15]]
    for file in files:
        if file in test:
            t=0
            path, ext = os.path.splitext(file)
            sub_df = pd.read_csv(file)
            df_list_test.append(sub_df)
        else :
            t=0
            path, ext = os.path.splitext(file)
            sub_df = pd.read_csv(file)
            df_list_train.append(sub_df)

    df_train2 = pd.concat(df_list_train)
    df_test2 = pd.concat(df_list_test)


    print(1)
    df_train2.columns

    print(df_train2)
    print(df_test2)

    #パラメータ保存用
    C=[]
    kernel= []
    gamma=[]
    #クラス間でのデータ数の不均衡を考慮
    #https://qiita.com/kento1109/items/1fc7488163b0f350f2fa
    poms=["workload_Future"]
    for i in poms:
        #print(i+"モデル")
        #テストデータと訓練データを読み取り
        #train_dataset=pd.read_csv("train_dataset_onehot_std_cla_10T.csv",index_col=0)
        #test_dataset=pd.read_csv("test_dataset_onehot_std_cla_10T.csv",index_col=0)
        
        #test_dataset=test_dataset.reset_index(drop=True)
        #train_dataset=train_dataset.reset_index(drop=True)
        
        #使わない列を削除
        #df_train2=train_dataset.drop(["id","height","weight","T-A","D-D","A-H","V","F","C","less_1cup","1cup_3cups","3cups_6cups","6cups_9cups","delta"], axis=1)
        #df_test2=test_dataset.drop(["id","height","weight","T-A","D-D","A-H","V","F","C","less_1cup","1cup_3cups","3cups_6cups","6cups_9cups","delta"], axis=1)

        
        #特徴量
        names=df_train2.columns
        
        #目的変数と説明変数に分ける
        x_train = df_train2.drop(i, axis=1).values
        y_train= df_train2[i].values
        x_test=df_test2.drop(i, axis=1).values
        y_test= df_test2[i].values
        
        #データを整形する
        Xtrain2 = np.array(x_train)
        Xtest2 = np.array(x_test)
        y_train2 = np.array(y_train)
        y_test2 = np.array(y_test)
        
        print(Xtrain2)
        
        #パラメータ
        tuned_parameters = [
            #{'C': [10], 'kernel': ['linear']},
            {'C': [1, 10, 100], 'kernel': ['rbf'], 'gamma': [0.01, 0.001, 0.0001]}
        ]
        
        print("パラメータ：",tuned_parameters)
        
        #最適化の実行
        score = 'accuracy'
        clf = GridSearchCV(
            SVC(class_weight='balanced'), # 識別器
            tuned_parameters, # 最適化したいパラメータセット 
            cv=5, # 交差検定の回数
            scoring='accuracy' ) # モデルの評価関数の指定
            #scoring='%s_weighted' % score ) # モデルの評価関数の指定
        print("clf",clf)
        clf.fit(Xtrain2, y_train2)

        #最適なモデルの選択-------------------------------------------------------------------------------------------------------------------------------
        best_model = clf.best_estimator_
        best_model_params=clf.best_params_
        print("最適なモデルのパラメータ",best_model_params)
        
        C_ = best_model_params["C"]
        kernel_ = best_model_params["kernel"]
        C.append(best_model_params["C"])
        kernel.append(best_model_params["kernel"])
        if kernel_=="rbf":
            gamma_=best_model_params["gamma"]
            gamma.append(best_model_params["gamma"])
            #モデルの構築
            model2=SVC(C=C_,kernel=kernel_,gamma=gamma_,class_weight='balanced',random_state=0)
        
        else:
            model2 = SVC(C=C_,kernel=kernel_,class_weight='balanced',random_state=0)
        print("model2",model2)
        
        # モデルのコンパイル
        #model_.compile(loss='mse',
        #              optimizer=tf.keras.optimizers.Adam(learning_rate=lr_),
        #              metrics=['mae', 'mse'])

        # モデルの学習
        history_ = model2.fit(Xtrain2, y_train2)
        print("モデルの学習",history_)
        #--------------------------------------------------------------------------------------------------------------------------------------------------
        
        #各試行でのスコアを確認
        #print(clf.best_estimator_)#最も性能がよかったランダムフォレストのインスタンス
        #print('Best params: {}'.format(clf.best_params_)) 
        #print('Best Score: {}'.format(clf.best_score_))#'criterion': 'entropy'の場合は小さいほうが良い、gridの場合は大きいほうがよい
        
        # 予測値算出
        #https://aiacademy.jp/media/?p=258
        pred_train = model2.predict(Xtrain2)
        print("グリッドサーチ・SVC")
        print( "\n [ 訓練データ結果 ]" )
        print("適合率（Precision）, 再現率（Recall）, F値(f1-scoreのavg/totalの部分)")
        train_class = classification_report(y_train2, pred_train,output_dict = True)
        print( train_class )


        print( "\n [ 混同行列 ]" )
        train_conf = confusion_matrix(y_train2, pred_train)
        print( train_conf )

        print( "\n [ 正解率 ]" )#予測結果全体がどれくらい真の値と一致しているかを表す指標
        print( accuracy_score(y_train2, pred_train) )
        
        
        # 予測値算出
        #https://aiacademy.jp/media/?p=258
        pred_test= model2.predict(Xtest2)
        print("グリッドサーチ・SVC")
        print( "\n [ テストデータ結果 ]" )
        print("適合率（Precision）, 再現率（Recall）, F値(f1-scoreのavg/totalの部分)")
        test_class = classification_report(y_test2, pred_test,output_dict = True)
        print( test_class )
        y_test2 = pd.DataFrame(y_test2)
        pred_test = pd.DataFrame(pred_test)
        y_test2.to_csv('C:\\Users\\rikua\\Documents\\pic_modify2\\SVM_test.csv')
        pred_test.to_csv('C:\\Users\\rikua\\Documents\\pic_modify2\\SVM_pred.csv')

        """
        test_class2 = []
        test_class3 = []
        for j in range(len(test_class)):
            for k in range(len(test_class.columns)):
                test_class2.append(test_class.iat[j,k])
            test_class3.append(test_class2)

        for j in range(len(test_class3)):
            for k in range(len(test_class3[1])):
                ws.cell(j+1,k+10,value = test_class3[j][k])"""

        print( "\n [ 混同行列 ]" )
        test_conf = confusion_matrix(y_test2, pred_test)
        print( test_conf )

        print( "\n [ 正解率 ]" )#予測結果全体がどれくらい真の値と一致しているかを表す指標
        print( accuracy_score(y_test2, pred_test) )

        
        report_df = pd.DataFrame(train_class).T
        report_df.to_csv('C:\\Users\\rikua\\Documents\\pic_modify2\\svm_train'+str(l)+'.csv')
        report_df = pd.DataFrame(test_class).T
        report_df.to_csv('C:\\Users\\rikua\\Documents\\pic_modify2\\svm_test'+str(l)+'.csv')
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]

        for j in range(len(train_conf)):
            for k in range(len(train_conf[1])):
                ws.cell(j+1,k+1,value = train_conf[j][k])
                ws.cell(j+7,k+1,value = test_conf[j][k])
        
        #wb.save('C:\\Users\\rikua\\Documents\\pic_modify2\\svm'+ str(l) + '.xlsx')

