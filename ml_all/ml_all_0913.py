import pandas as pd#pandasのインポート
import datetime#元データの日付処理のためにインポート
import numpy as np
from sklearn.model_selection import train_test_split, GridSearchCV#データ分割用
from sklearn.metrics import classification_report, accuracy_score, confusion_matrix,precision_score,recall_score,f1_score
from sklearn.metrics import r2_score#決定係数求める用
import matplotlib.pyplot as plt
#%matplotlib inline
import joblib #モデルの保存、ロード用
import pickle#モデルの保存、ロード用
from pandas import Series, DataFrame
import sklearn
import pathlib
from sklearn import datasets, preprocessing
from sklearn.model_selection import GridSearchCV
from sklearn import *
from collections import Counter
from imblearn.over_sampling import SMOTE
from sklearn.neural_network import MLPClassifier
import glob
import random
import os
from sklearn.metrics import f1_score
from sklearn.metrics import make_scorer
import openpyxl
import pandas as pd#pandasのインポート
import datetime#元データの日付処理のためにインポート
from sklearn import ensemble, model_selection
from sklearn.model_selection import train_test_split, GridSearchCV#データ分割用
from sklearn.model_selection import RandomizedSearchCV#ランダムサーチ
from sklearn.ensemble import RandomForestClassifier as RFC#ランダムフォレスト(クラス分類用)
from sklearn.model_selection import learning_curve#学習曲線
from sklearn.model_selection import RandomizedSearchCV#ランダムサーチ
from sklearn.model_selection import cross_val_score # クロスバリデーション用
from sklearn.svm import SVC # SVM用
from sklearn import metrics # 精度検証用
from sklearn.metrics import r2_score#決定係数求める用
import matplotlib.pyplot as plt
from sklearn.model_selection import train_test_split
import scipy.stats
import smtplib


def NNclass():

    poms=["workload"]


    for i in poms:
        search_params = {
        #'hidden_layer_sizes'      : [(50,50)],#隠れ層のノード数
        #'hidden_layer_sizes'      : [(50,50),(100,100),(500,500),(1000,1000)],#隠れ層のノード数
        'hidden_layer_sizes'      : [(50,50)],#隠れ層のノード数
        'activation': ["relu"],#活性化関数
        'solver'      : ['adam'],#最適化手法(大きなデータセットだとadam,小さなデータセットではlbfgs)
        'max_iter'            : [300,500,700,800,900,1000],#最大エポック数
        # 'max_iter'            : [900],#最大エポック数
        #'max_iter'            : [900],#最大エポック数
        }
        #パラメータ
        #search_params = [
        #    {'hidden_layer_sizes' : [10,100,1000], 'activation': ["relu", "tanh"],'solver' : ['lbfgs'],'max_iter':[1000] },
        #    {'hidden_layer_sizes' : [10,100,1000], 'activation': ["relu", "tanh"], 'solver': ["adam","sgd"],'batch_size':[64,128,256,512],'max_iter':[1000]}
        #]
        
        clf = GridSearchCV(estimator=MLPClassifier(random_state=0),param_grid=search_params, scoring="f1_macro" , cv=5,verbose=True,n_jobs=-1) 
        
        #clf = MLPClassifier(hidden_layer_sizes=10, activation='relu',solver='adam', max_iter=1000,random_state=0)
        clf.fit(train_x, train_y)
        #print('accuracy_score: %.3f' % clf.score(test_x, test_y))
        best_model = clf.best_estimator_
        best_model_params=clf.best_params_
        
        hidden_layer_sizes_ = best_model_params["hidden_layer_sizes"]
        activation_ = best_model_params["activation"]
        solver_=best_model_params["solver"]
        max_iter_ = best_model_params["max_iter"]
        
        print(best_model_params)
        
        #モデルの構築
        #model2 = RFC(criterion=criterion_, max_depth = max_depth_, max_features=max_features_ , n_estimators=n_estimators_,random_state=0,verbose=1)
        model2 = MLPClassifier(hidden_layer_sizes=hidden_layer_sizes_,activation=activation_,solver=solver_,early_stopping = True,  shuffle = False,max_iter=max_iter_,random_state=0,verbose=True)
        
        print(1)

        model2.fit(train_x, train_y)
        print("######################################")

        # with open('C:\\Users\\naklab\\Desktop\\ml_all\\NNmodeltwo0827.pickle', mode='wb') as f:  # with構文でファイルパスとバイナリ書き込みモードを設定
        #     pickle.dump(model2, f)  

        print("iine")
        print("######################################")

        
        pred_train =model2.predict(train_x)
        print("NNT"+i+"モデル")
        print( "\n [ 訓練データ結果 ]" )
        print("適合率（Precision）, 再現率（Recall）, F値(f1-scoreのavg/totalの部分)")
        train_class = classification_report(train_y, pred_train,output_dict = True)
        print( train_class )

        print( "\n [ 混同行列 ]" )
        train_conf = confusion_matrix(train_y, pred_train)
        print( train_conf )

        print( "\n [ 正解率 ]" )#予測結果全体がどれくらい真の値と一致しているかを表す指標
        print( accuracy_score(train_y, pred_train) )
        
        pred_test = model2.predict(test_x)
        print("NNT"+i+"モデル")
        print( "\n [ テストデータ結果 ]" )
        print("適合率（Precision）, 再現率（Recall）, F値(f1-scoreのavg/totalの部分)")
        test_class = classification_report(test_y, pred_test,output_dict = True)
        print( test_class )

        print( "\n [ 混同行列 ]" )
        test_conf = confusion_matrix(test_y, pred_test) 
        print( test_conf )



        print( "\n [ 正解率 ]" )#予測結果全体がどれくらい真の値と一致しているかを表す指標
        print( accuracy_score(test_y, pred_test) )
        plt.plot(model2.loss_curve_)
        plt.title("Loss Curve", fontsize=14)
        plt.xlabel('Iterations')
        plt.ylabel('Cost')
        #plt.show()

        


        report_df = pd.DataFrame(train_class).T
        report_df.to_csv(path + 'nn_train'+str(l)+'.csv')
        report_df = pd.DataFrame(test_class).T
        report_df.to_csv(path + 'nn_test'+str(l)+'.csv')
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]

        for j in range(len(train_conf)):
            for k in range(len(train_conf[1])):
                ws.cell(j+1,k+1,value = train_conf[j][k])
                ws.cell(j+7,k+1,value = test_conf[j][k])
        
        wb.save(path + 'nn'+ str(l) + '.xlsx')

    

def RDFclass():
    criterion_array=[]
    max_depth_array=[]
    max_features_array=[]
    n_estimators_array =[]

    poms=["workload"]
    #グリッドサーチ、分類のバランスを考慮、ラベルエンコーディングしたとき
    #https://qiita.com/kazuki_hayakawa/items/6d7a4597829f54ebdb83
    for i in poms:
        
        search_params = {
        'criterion': ['gini', 'entropy'],
        'n_estimators'      : [10,20,50,100],#決定木の数
        'max_features'      : ['sqrt', 'log2'],#個々の決定木に使用する特徴量の数
        #'n_jobs'            : [1],
        #'min_samples_split' : [3, 5, 10, 20, 30, 40, 50, 100],
        #'min_samples_leaf': [5,50],
        'max_depth'         : [5,10,20,50,100]#決定木最大の深さ
        }
        """
        search_params = {
        'criterion': ['gini', 'entropy'],
        'n_estimators'      : [i for i in range(1,10)],#決定木の数
        'max_features'      : ['sqrt', 'log2'],#個々の決定木に使用する特徴量の数
        'max_depth'         : [2,3,4,5,6,7]#決定木最大の深さ
        }
        """
        RFC_grid = GridSearchCV(estimator=RFC(random_state=0,class_weight='balanced'),param_grid=search_params, scoring='accuracy', cv=5,verbose=True,n_jobs=-1) #グリッドサーチ・ランダムフォレスト
        
        RFC_grid.fit(train_x, train_y)
        best_model = RFC_grid.best_estimator_
        best_model_params=RFC_grid.best_params_
        
        criterion_ = best_model_params["criterion"]
        max_depth_ = best_model_params["max_depth"]
        max_features_=best_model_params["max_features"]
        #min_samples_split_ = best_model_params["min_samples_split"]
        #min_samples_leaf_=best_model_params["min_samples_leaf"]
        n_estimators_ = best_model_params["n_estimators"]
        #n_jobs_=best_model_params["n_jobs"]
        
        #パラメータを配列にいれる
        criterion_array.append(criterion_)
        max_depth_array.append(max_depth_)
        max_features_array.append(max_features_)
        #min_samples_split_array.append(min_samples_split_)
        #min_samples_leaf_array.append(min_samples_leaf_)
        n_estimators_array .append(n_estimators_)
        #n_jobs_array.append(n_jobs_)
        
        #モデルの構築
        model2 = RFC(class_weight='balanced',criterion=criterion_, max_depth = max_depth_, max_features=max_features_ , n_estimators=n_estimators_,random_state=0,verbose=1)
        #model2 = RFC(class_weight='balanced',criterion=criterion_, max_depth = max_depth_, max_features=max_features_,min_samples_split= min_samples_split_,n_estimators=n_estimators_,n_jobs=n_jobs_,random_state=0,verbose=1)

        # モデルのコンパイル
        #model_.compile(loss='mse',
        #              optimizer=tf.keras.optimizers.Adam(learning_rate=lr_),
        #              metrics=['mae', 'mse'])

        # モデルの学習
        history_ = model2.fit(train_x, train_y)

        # モデルの保存
        with open('C:\\Users\\naklab\\Desktop\\ml_all\\RDFmodeltwo0825.pickle', mode='wb') as f:  # with構文でファイルパスとバイナリ書き込みモードを設定
            pickle.dump(history_, f)                   # オブジェクトをシリアライズ
        
        # モデルを保存する(https://localab.jp/blog/save-and-load-machine-learning-models-in-python-with-scikit-learn/)
        #filename = 'RDF_Classification_grid'+i+'finalized_model.sav'
        #joblib.dump(history_, open(filename, 'wb'))
        
        pred_train = history_.predict(train_x)
        
        # 予測値算出
        #https://aiacademy.jp/media/?p=258
        pred_train = model2.predict(train_x)
        print("グリッドサーチ・ランダムフォレスト")
        print( "\n [ 訓練データ結果 ]" )
        print("適合率（Precision）, 再現率（Recall）, F値(f1-scoreのavg/totalの部分)")
        train_class = classification_report(train_y, pred_train,output_dict = True)
        print( train_class )

        print( "\n [ 混同行列 ]" )
        train_conf = confusion_matrix(train_y, pred_train)
        print( train_conf )

        print( "\n [ 正解率 ]" )#予測結果全体がどれくらい真の値と一致しているかを表す指標
        print( accuracy_score(train_y, pred_train) )
        
        print(i+'モデルのグリッドサーチ・ランダムフォレストモデルにおける n_estimators   :  %d'  %RFC_grid.best_estimator_.n_estimators)
        
        print("グリッドサーチ・ランダムフォレスト")
        print(RFC_grid.best_estimator_)#最も性能がよかったランダムフォレストのインスタンス
        print('Best params: {}'.format(RFC_grid.best_params_)) 
        print('Best Score: {}'.format(RFC_grid.best_score_))#'criterion': 'entropy'の場合は小さいほうが良い、gridの場合は大きいほうがよい

        # 予測値算出
        #https://aiacademy.jp/media/?p=258
        pred_test = model2.predict(test_x)
        print( "\n [ テストデータ結果 ]" )
        print("適合率（Precision）, 再現率（Recall）, F値(f1-scoreのavg/totalの部分)")
        test_class = classification_report(test_y, pred_test,output_dict = True)
        print( test_class )

        print( "\n [ 混同行列 ]" )
        test_conf = confusion_matrix(test_y, pred_test)
        print( test_conf )

        print( "\n [ 正解率 ]" )#予測結果全体がどれくらい真の値と一致しているかを表す指標
        print( accuracy_score(test_y, pred_test) )
        
        feature = model2.feature_importances_
        # 特徴量の名前ラベルを取得
        label = df_list_train.columns[0:]
        # 特徴量の重要度順（降順）に並べて表示
        indices = np.argsort(feature)[::-1]
        for i in range(len(feature)):
            print(str(i + 1) + "   " +
                str(label[indices[i]]) + "   " + str(feature[indices[i]]))
            
        # 実際の値と予測値の比較グラフ
        plt.subplot(121, facecolor='white')
        #plt_label = [i for i in range(1, 32)]
        plt.plot(test_y, color='blue')
        plt.plot(pred_test, color='red')
        # 特徴量の重要度の棒グラフ
        plt.subplot(122, facecolor='white')
        plt.title('特徴量の重要度')
        plt.bar(
            range(
                len(feature)),
            feature[indices],
            color='blue',
            align='center')
        plt.xticks(range(len(feature)), label[indices], rotation=45)
        plt.xlim([-1, len(feature)])
        plt.tight_layout()
        # グラフの表示
        #plt.show()

                
        report_df = pd.DataFrame(train_class).T
        report_df.to_csv(path + 'rdf_train'+str(l)+'.csv')
        report_df = pd.DataFrame(test_class).T
        report_df.to_csv(path + 'rdf_test'+str(l)+'.csv')
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]

        for j in range(len(train_conf)):
            for k in range(len(train_conf[1])):
                ws.cell(j+1,k+1,value = train_conf[j][k])
                ws.cell(j+7,k+1,value = test_conf[j][k])
        
        wb.save(path + 'rdf'+ str(l) + '.xlsx')


def SVMclass():
    #パラメータ保存用
    C=[]
    kernel= []
    gamma=[]
    #クラス間でのデータ数の不均衡を考慮
    #https://qiita.com/kento1109/items/1fc7488163b0f350f2fa
    poms=["workload"]
    for i in poms:
        #print(i+"モデル")
        #テストデータと訓練データを読み取り
        #train_dataset=pd.read_csv("train_dataset_onehot_std_cla_10T.csv",index_col=0)
        #test_dataset=pd.read_csv("test_dataset_onehot_std_cla_10T.csv",index_col=0)
        
        #test_dataset=test_dataset.reset_index(drop=True)
        #train_dataset=train_dataset.reset_index(drop=True)
        
        #使わない列を削除
        #df_train2=train_dataset.drop(["id","height","weight","T-A","D-D","A-H","V","F","C","less_1cup","1cup_3cups","3cups_6cups","6cups_9cups","delta"], axis=1)
        #df_test2=test_dataset.drop(["id","height","weight","T-A","D-D","A-H","V","F","C","less_1cup","1cup_3cups","3cups_6cups","6cups_9cups","delta"], axis=1)

        
        #パラメータ
        tuned_parameters = [
            #{'C': [10], 'kernel': ['linear']},
            {'C': [1, 10, 100], 'kernel': ['rbf'], 'gamma': [0.01, 0.001, 0.0001]}
        ]
        
        print("パラメータ：",tuned_parameters)
        
        #最適化の実行
        score = 'accuracy'
        clf = GridSearchCV(
            SVC(class_weight='balanced'), # 識別器
            tuned_parameters, # 最適化したいパラメータセット 
            cv=5, # 交差検定の回数
            scoring='accuracy' ) # モデルの評価関数の指定
            #scoring='%s_weighted' % score ) # モデルの評価関数の指定
        print("clf",clf)
        clf.fit(train_x, train_y)

        #最適なモデルの選択-------------------------------------------------------------------------------------------------------------------------------
        best_model = clf.best_estimator_
        best_model_params=clf.best_params_
        print("最適なモデルのパラメータ",best_model_params)
        
        C_ = best_model_params["C"]
        kernel_ = best_model_params["kernel"]
        C.append(best_model_params["C"])
        kernel.append(best_model_params["kernel"])
        if kernel_=="rbf":
            gamma_=best_model_params["gamma"]
            gamma.append(best_model_params["gamma"])
            #モデルの構築
            model2=SVC(C=C_,kernel=kernel_,gamma=gamma_,class_weight='balanced',random_state=0)
        
        else:
            model2 = SVC(C=C_,kernel=kernel_,class_weight='balanced',random_state=0)
        print("model2",model2)
        
        # モデルのコンパイル
        #model_.compile(loss='mse',
        #              optimizer=tf.keras.optimizers.Adam(learning_rate=lr_),
        #              metrics=['mae', 'mse'])

        # モデルの学習
        history_ = model2.fit(train_x, train_y)
        print("モデルの学習",history_)
        #--------------------------------------------------------------------------------------------------------------------------------------------------
        
        #各試行でのスコアを確認
        #print(clf.best_estimator_)#最も性能がよかったランダムフォレストのインスタンス
        #print('Best params: {}'.format(clf.best_params_)) 
        #print('Best Score: {}'.format(clf.best_score_))#'criterion': 'entropy'の場合は小さいほうが良い、gridの場合は大きいほうがよい
        
        # 予測値算出
        #https://aiacademy.jp/media/?p=258
        pred_train = model2.predict(train_x)
        print("グリッドサーチ・SVC")
        print( "\n [ 訓練データ結果 ]" )
        print("適合率（Precision）, 再現率（Recall）, F値(f1-scoreのavg/totalの部分)")
        train_class = classification_report(train_y, pred_train,output_dict = True)
        print( train_class )


        print( "\n [ 混同行列 ]" )
        train_conf = confusion_matrix(train_y, pred_train)
        print( train_conf )

        print( "\n [ 正解率 ]" )#予測結果全体がどれくらい真の値と一致しているかを表す指標
        print( accuracy_score(train_y, pred_train) )
        
        
        # 予測値算出
        #https://aiacademy.jp/media/?p=258
        pred_test= model2.predict(test_x)
        print("グリッドサーチ・SVC")
        print( "\n [ テストデータ結果 ]" )
        print("適合率（Precision）, 再現率（Recall）, F値(f1-scoreのavg/totalの部分)")
        test_class = classification_report(test_y, pred_test,output_dict = True)
        print( test_class )

        """
        test_class2 = []
        test_class3 = []
        for j in range(len(test_class)):
            for k in range(len(test_class.columns)):
                test_class2.append(test_class.iat[j,k])
            test_class3.append(test_class2)

        for j in range(len(test_class3)):
            for k in range(len(test_class3[1])):
                ws.cell(j+1,k+10,value = test_class3[j][k])"""

        print( "\n [ 混同行列 ]" )
        test_conf = confusion_matrix(test_y, pred_test)
        print( test_conf )

        print( "\n [ 正解率 ]" )#予測結果全体がどれくらい真の値と一致しているかを表す指標
        print( accuracy_score(test_y, pred_test) )

        
        report_df = pd.DataFrame(train_class).T
        report_df.to_csv(path + 'svm_train'+str(l)+'.csv')
        report_df = pd.DataFrame(test_class).T
        report_df.to_csv(path + 'svm_test'+str(l)+'.csv')
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]

        for j in range(len(train_conf)):
            for k in range(len(train_conf[1])):
                ws.cell(j+1,k+1,value = train_conf[j][k])
                ws.cell(j+7,k+1,value = test_conf[j][k])
        
        wb.save(path + 'svm'+ str(l) + '.xlsx')


l_list = [0, 300, 600, 900]

for l in l_list:
    df_list_test = []
    df_list_train = []
    files = glob.glob('now_all/*')
    test = [files[1],files[5],files[7],files[9],files[11],files[14],files[17],files[19],files[22],files[24]]
    path ="output/"
    for file in files:
        if file in test:
            sub_df = pd.read_csv(file)
            sub_wl = sub_df.pop("workload")
            sub_wl = sub_wl[l:]
            sub_wl = sub_wl.reset_index(drop=True)
            sub_df = pd.concat([sub_df, sub_wl],axis = 1)
            sub_df = sub_df.dropna()
            df_list_test.append(sub_df)
        else :
            sub_df = pd.read_csv(file)
            sub_wl = sub_df.pop("workload")
            sub_wl = sub_wl[l:]
            sub_wl = sub_wl.reset_index(drop=True)
            sub_df = pd.concat([sub_df, sub_wl],axis = 1)
            sub_df = sub_df.dropna()
            df_list_train.append(sub_df)
    
    df_list_train = pd.concat(df_list_train)
    df_list_test = pd.concat(df_list_test)
    print(df_list_train)
    print(df_list_test)

    train_x = df_list_train.drop("workload",axis=1)
    train_y= df_list_train[["workload"]]
    test_x=df_list_test.drop("workload",axis=1)
    test_y= df_list_test[["workload"]]

    sm = SMOTE()

    print(train_x)
    print(test_x)

    train_x, train_y = sm.fit_resample(train_x, train_y)
    test_x, test_y = sm.fit_resample(test_x, test_y)
    
    print("________________________")

    print(train_x)
    print(test_x)


    train3 = []
    test3 = []




    train2 = pd.concat([train_x,train_y], axis=1)
    test2 = pd.concat([test_x, test_y], axis=1)
    print(train2)
    print(test2)
    for j in range(1,6):

        train4 = train2[train2['workload'] == j]
        test4 = test2[test2['workload'] == j]
        train4 = train4[::5]
        test4 = test4[::5]    
        train3.append(train4)
        test3.append(test4)
        print(len(train4))
    train3 = pd.concat(train3)
    test3 = pd.concat(test3)
    print("________________________")

    print(train3)
    print(test3)

    train_x2 = train3.drop("workload",axis=1)
    train_y2= train3[["workload"]]
    test_x2=test3.drop("workload",axis=1)
    test_y2= test3[["workload"]]       



    train_x = np.array(train_x2)

    test_x = np.array(test_x2)
    train_y = np.array(train_y2)
    test_y = np.array(test_y2)


    NNclass()
    RDFclass()
    SVMclass()

    smtp_host = 'smtp.gmail.com'
    smtp_port = 465
    username = 'kalashnikova1120@gmail.com'
    password = 'nfjaizuqmrpspvti'
    from_address = 'kalashnikova1120@gmail.com'
    to_address = 'kalashnikova1120@gmail.com'
    subject = 'test subject'
    body = 'test body'
    message = ("From: %srnTo: %srnSubject: %srnrn%s" % (from_address, to_address, subject, body))

    smtp = smtplib.SMTP_SSL(smtp_host, smtp_port)
    smtp.login(username, password)
    result = smtp.sendmail(from_address, to_address, message)
    print(result)
