import pandas as pd#pandasのインポート
import datetime#元データの日付処理のためにインポート
import numpy as np
from sklearn import ensemble, model_selection
from sklearn.model_selection import train_test_split, GridSearchCV#データ分割用
from sklearn.metrics import classification_report, accuracy_score, confusion_matrix,precision_score,recall_score,f1_score
from sklearn.metrics import r2_score#決定係数求める用
import matplotlib.pyplot as plt
#%matplotlib inline
import joblib#モデルの保存、ロード用
import pickle#モデルの保存、ロード用
from pandas import Series, DataFrame
import sklearn
import pathlib
from sklearn import datasets, preprocessing
from sklearn.model_selection import GridSearchCV
from sklearn import *
from collections import Counter
from imblearn.over_sampling import SMOTE
from sklearn.neural_network import MLPClassifier
import glob
import random
import os
from sklearn.metrics import f1_score
from sklearn.metrics import make_scorer
import openpyxl

for l in range(3):
    if l == 0:
        files = glob.glob('C:\\Users\\rikua\\Documents\\all_csv_mini_five\\*')
    elif l == 1:
        files = glob.glob('C:\\Users\\rikua\\Documents\\all_csv_mini_five\\*')
    else :
        files = glob.glob('C:\\Users\\rikua\\Documents\\all_csv_mini_five\\*')

    f1_scorer = make_scorer(f1_score) 
    df_list_test = []
    df_list_train = []
    test = random.sample(files,4)
    test = [files[0],files[5],files[11],files[15]]
    for file in files:
        if file in test:
            t=0
            path, ext = os.path.splitext(file)
            sub_df = pd.read_csv(file)
            df_list_test.append(sub_df)
        else :
            t=0
            path, ext = os.path.splitext(file)
            sub_df = pd.read_csv(file)
            df_list_train.append(sub_df)

    df_train2 = pd.concat(df_list_train)
    df_test2 = pd.concat(df_list_test)


    print(1)
    df_train2.columns

    print(df_train2)
    print(df_test2)

    poms=["workload"]

    for i in poms:
        search_params = {
        #'hidden_layer_sizes'      : [(50,50)],#隠れ層のノード数
        'hidden_layer_sizes'      : [(50,50),(100,100),(500,500),(1000,1000)],#隠れ層のノード数
        'activation': ["relu"],#活性化関数
        'solver'      : ['adam'],#最適化手法(大きなデータセットだとadam,小さなデータセットではlbfgs)
        'max_iter'            : [300,500,700,800,900,1000],#最大エポック数
        #'max_iter'            : [900],#最大エポック数
        }

        train_x = df_train2.drop(i, axis=1).values
        train_y= df_train2[i].values
        test_x=df_test2.drop(i, axis=1).values
        test_y= df_test2[i].values

        train_x = np.array(train_x)
        test_x = np.array(test_x)
        train_y = np.array(train_y)
        test_y = np.array(test_y)
        

    
        #パラメータ
        #search_params = [
        #    {'hidden_layer_sizes' : [10,100,1000], 'activation': ["relu", "tanh"],'solver' : ['lbfgs'],'max_iter':[1000] },
        #    {'hidden_layer_sizes' : [10,100,1000], 'activation': ["relu", "tanh"], 'solver': ["adam","sgd"],'batch_size':[64,128,256,512],'max_iter':[1000]}
        #]
        
        clf = GridSearchCV(estimator=MLPClassifier(random_state=0),param_grid=search_params, scoring="f1_macro" , cv=5,verbose=True,n_jobs=-1) 
        
        #clf = MLPClassifier(hidden_layer_sizes=10, activation='relu',solver='adam', max_iter=1000,random_state=0)
        clf.fit(train_x, train_y)
        #print('accuracy_score: %.3f' % clf.score(test_x, test_y))
        best_model = clf.best_estimator_
        best_model_params=clf.best_params_
        
        hidden_layer_sizes_ = best_model_params["hidden_layer_sizes"]
        activation_ = best_model_params["activation"]
        solver_=best_model_params["solver"]
        max_iter_ = best_model_params["max_iter"]
        
        print(best_model_params)
        
        #モデルの構築
        #model2 = RFC(criterion=criterion_, max_depth = max_depth_, max_features=max_features_ , n_estimators=n_estimators_,random_state=0,verbose=1)
        model2 = MLPClassifier(hidden_layer_sizes=hidden_layer_sizes_,activation=activation_,solver=solver_,early_stopping = True,  shuffle = False,max_iter=max_iter_,random_state=0,verbose=True)
        
        print(1)

        model2.fit(train_x, train_y)

        print("iine")
        
        pred_train =model2.predict(train_x)
        print("NNT"+i+"モデル")
        print( "\n [ 訓練データ結果 ]" )
        print("適合率（Precision）, 再現率（Recall）, F値(f1-scoreのavg/totalの部分)")
        train_class = classification_report(train_y, pred_train,output_dict = True)
        print( train_class )

        print( "\n [ 混同行列 ]" )
        train_conf = confusion_matrix(train_y, pred_train)
        print( train_conf )

        print( "\n [ 正解率 ]" )#予測結果全体がどれくらい真の値と一致しているかを表す指標
        print( accuracy_score(train_y, pred_train) )
        
        pred_test = model2.predict(test_x)
        print("NNT"+i+"モデル")
        print( "\n [ テストデータ結果 ]" )
        print("適合率（Precision）, 再現率（Recall）, F値(f1-scoreのavg/totalの部分)")
        test_class = classification_report(test_y, pred_test,output_dict = True)
        print( test_class )

        print( "\n [ 混同行列 ]" )
        test_conf = confusion_matrix(test_y, pred_test) 
        print( test_conf )



        print( "\n [ 正解率 ]" )#予測結果全体がどれくらい真の値と一致しているかを表す指標
        print( accuracy_score(test_y, pred_test) )
        plt.plot(model2.loss_curve_)
        plt.title("Loss Curve", fontsize=14)
        plt.xlabel('Iterations')
        plt.ylabel('Cost')
        #plt.show()

        


        report_df = pd.DataFrame(train_class).T
        report_df.to_csv('C:\\Users\\rikua\\Documents\\pic_modify2\\nn_train'+str(l)+'.csv')
        report_df = pd.DataFrame(test_class).T
        report_df.to_csv('C:\\Users\\rikua\\Documents\\pic_modify2\\nn_test'+str(l)+'.csv')
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]

        for j in range(len(train_conf)):
            for k in range(len(train_conf[1])):
                ws.cell(j+1,k+1,value = train_conf[j][k])
                ws.cell(j+7,k+1,value = test_conf[j][k])
        
        wb.save('C:\\Users\\rikua\\Documents\\pic_modify2\\nn'+ str(l) + '.xlsx')

    
