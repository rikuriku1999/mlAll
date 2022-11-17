import pandas as pd#pandasのインポート
import datetime#元データの日付処理のためにインポート
import numpy as np
from sklearn.model_selection import train_test_split, GridSearchCV#データ分割用
from sklearn.metrics import classification_report, accuracy_score, confusion_matrix,precision_score,recall_score,f1_score
from sklearn.metrics import r2_score#決定係数求める用
import matplotlib.pyplot as plt
#%matplotlib inline
import joblib #モデルの保存、ロード用
import pickle#モデルの保存、ロード用
from pandas import Series, DataFrame
import sklearn
import pathlib
from sklearn import datasets, preprocessing
from sklearn.model_selection import GridSearchCV
from sklearn import *
from collections import Counter
from imblearn.over_sampling import SMOTE
from sklearn.neural_network import MLPClassifier
import glob
import random
import os
from sklearn.metrics import f1_score
from sklearn.metrics import make_scorer
import openpyxl
import pandas as pd#pandasのインポート
import datetime#元データの日付処理のためにインポート
from sklearn import ensemble, model_selection
from sklearn.model_selection import train_test_split, GridSearchCV#データ分割用
from sklearn.model_selection import RandomizedSearchCV #ランダムサーチ
from sklearn.ensemble import RandomForestClassifier as RFC #ランダムフォレスト(クラス分類用)
from sklearn.model_selection import learning_curve #学習曲線
from sklearn.model_selection import RandomizedSearchCV#ランダムサーチ
from sklearn.model_selection import cross_val_score # クロスバリデーション用
from sklearn.svm import SVC # SVM用
from sklearn import metrics # 精度検証用
from sklearn.metrics import r2_score#決定係数求める用
import matplotlib.pyplot as plt
from sklearn.model_selection import train_test_split
import scipy.stats
import smtplib


def NNclass(
        train_x=None,
        train_y=None,
        test_x=None,
        test_y=None,
        title = "hoge",
        hidden_layer_sizes = 300, 
        solver = "adam", 
        activation = "relu", 
        max_iter = 300
        ):
    poms=["workload"]

    for i in poms:
        search_params = [
           {
            'hidden_layer_sizes' : hidden_layer_sizes, 
            'activation': activation,
            'solver' : solver,
           },
        ]
             
        clf = GridSearchCV(estimator=MLPClassifier(random_state=0),param_grid=search_params, scoring="f1_macro" , cv=5,verbose=True,n_jobs=-1) 
        
        #clf = MLPClassifier(hidden_layer_sizes=10, activation='relu',solver='adam', max_iter=1000,random_state=0)
        clf.fit(train_x, train_y)
        #print('accuracy_score: %.3f' % clf.score(test_x, test_y))
        best_model = clf.best_estimator_
        best_model_params=clf.best_params_
        
        hidden_layer_sizes_ = best_model_params["hidden_layer_sizes"]
        activation_ = best_model_params["activation"]
        solver_=best_model_params["solver"]
        max_iter_ = best_model_params["max_iter"]
        
        print(best_model_params)
        
        #モデルの構築
        #model2 = RFC(criterion=criterion_, max_depth = max_depth_, max_features=max_features_ , n_estimators=n_estimators_,random_state=0,verbose=1)
        model2 = MLPClassifier(hidden_layer_sizes=hidden_layer_sizes_,activation=activation_,solver=solver_,early_stopping = True,  shuffle = False,max_iter=max_iter_,random_state=0,verbose=True)

        model2.fit(train_x, train_y)

        with open("models/" + 'nn'+ str(l) + '.pickle', mode='wb') as f:  # with構文でファイルパスとバイナリ書き込みモードを設定
            pickle.dump(model2, f)  

        pred_train =model2.predict(train_x)
        print("NNT"+i+"モデル")
        print( "\n [ 訓練データ結果 ]" )
        print("適合率（Precision）, 再現率（Recall）, F値(f1-scoreのavg/totalの部分)")
        train_class = classification_report(train_y, pred_train,output_dict = True)
        print( train_class )

        print( "\n [ 混同行列 ]" )
        train_conf = confusion_matrix(train_y, pred_train)
        print( train_conf )

        print( "\n [ 正解率 ]" )#予測結果全体がどれくらい真の値と一致しているかを表す指標
        print( accuracy_score(train_y, pred_train) )
        
        pred_test = model2.predict(test_x)
        print("NNT"+i+"モデル")
        print( "\n [ テストデータ結果 ]" )
        print("適合率（Precision）, 再現率（Recall）, F値(f1-scoreのavg/totalの部分)")
        test_class = classification_report(test_y, pred_test,output_dict = True)
        print( test_class )

        print( "\n [ 混同行列 ]" )
        test_conf = confusion_matrix(test_y, pred_test) 
        print( test_conf )



        print( "\n [ 正解率 ]" )#予測結果全体がどれくらい真の値と一致しているかを表す指標
        best_accuracy_score = accuracy_score(test_y, pred_test) 
        print( accuracy_score(test_y, pred_test) )

        report_df = pd.DataFrame(train_class).T
        report_df.to_csv("output/" + 'nn_train'+str(l)+'.csv')
        report_df = pd.DataFrame(test_class).T
        report_df.to_csv("output/" + 'nn_test'+str(l)+'.csv')
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]

        param_list = []
        param_list.append(["nn", str(l)])
        for key in best_model_params:
            param_list.append([key, best_model_params[key]])

        with open('parameter/' + 'parameter.txt', 'w') as o:
            for row in param_list:
                print(*row, sep=",", file=o)

        for j in range(len(train_conf)):
            for k in range(len(train_conf[1])):
                ws.cell(row = j+1,column = k+1,value = train_conf[j][k])
                ws.cell(row = j+7,column = k+1,value = test_conf[j][k])
        
        wb.save("output/" + 'nn'+ str(l) + title + '.xlsx')
    return best_model_params, test_conf, best_accuracy_score

    

def RDFclass(
            train_x=None,
            train_y=None,
            test_x=None,
            test_y=None,
            title="hoge", 
            criterion = "gini", 
            n_estimators = 100, 
            max_features = "sqrt", 
            min_samples_leaf = None, 
            min_samples_split = None, 
            max_depth = 30
            ):

    poms=["workload"]
    #グリッドサーチ、分類のバランスを考慮、ラベルエンコーディングしたとき
    #https://qiita.com/kazuki_hayakawa/items/6d7a4597829f54ebdb83
    for i in poms:
        search_params = {
        'criterion': criterion,
        'n_estimators'      : n_estimators,#決定木の数
        'max_features'      : max_features,#個々の決定木に使用する特徴量の数
        'max_depth'         : max_depth#決定木最大の深さ
        }  
        RFC_grid = GridSearchCV(estimator=RFC(random_state=0,class_weight='balanced'),param_grid=search_params, scoring='accuracy', cv=5,verbose=True,n_jobs=-1) #グリッドサーチ・ランダムフォレスト
        
        RFC_grid.fit(train_x, train_y)
        best_model = RFC_grid.best_estimator_
        best_model_params=RFC_grid.best_params_
        
        criterion_ = best_model_params["criterion"]
        max_depth_ = best_model_params["max_depth"]
        max_features_ = best_model_params["max_features"]
        min_samples_split_ = best_model_params["min_samples_split"]
        min_samples_leaf_ = best_model_params["min_samples_leaf"]
        n_estimators_ = best_model_params["n_estimators"]
        n_jobs_ = best_model_params["n_jobs"]
        
        #モデルの構築
        model2 = RFC(class_weight='balanced',criterion=criterion_, max_depth = max_depth_, max_features=max_features_ , n_estimators=n_estimators_,random_state=0,verbose=1)
        #model2 = RFC(class_weight='balanced',criterion=criterion_, max_depth = max_depth_, max_features=max_features_,min_samples_split= min_samples_split_,n_estimators=n_estimators_,n_jobs=n_jobs_,random_state=0,verbose=1)

        # モデルのコンパイル
        #model_.compile(loss='mse',
        #              optimizer=tf.keras.optimizers.Adam(learning_rate=lr_),
        #              metrics=['mae', 'mse'])

        # モデルの学習
        history_ = model2.fit(train_x, train_y)

        #モデルの保存
        with open("output/" + 'rdf'+ str(l) +'.pickle', mode='wb') as f:  # with構文でファイルパスとバイナリ書き込みモードを設定
            pickle.dump(history_, f)                   # オブジェクトをシリアライズ
        
        # モデルを保存する(https://localab.jp/blog/save-and-load-machine-learning-models-in-python-with-scikit-learn/)
        #filename = 'RDF_Classification_grid'+i+'finalized_model.sav'
        #joblib.dump(history_, open(filename, 'wb'))
        
        pred_train = history_.predict(train_x)
        
        # 予測値算出
        #https://aiacademy.jp/media/?p=258
        pred_train = model2.predict(train_x)
        print("グリッドサーチ・ランダムフォレスト")
        print( "\n [ 訓練データ結果 ]" )
        print("適合率（Precision）, 再現率（Recall）, F値(f1-scoreのavg/totalの部分)")
        train_class = classification_report(train_y, pred_train,output_dict = True)
        print( train_class )

        print( "\n [ 混同行列 ]" )
        train_conf = confusion_matrix(train_y, pred_train)
        print( train_conf )

        print( "\n [ 正解率 ]" )#予測結果全体がどれくらい真の値と一致しているかを表す指標
        print( accuracy_score(train_y, pred_train) )
        
        print(i+'モデルのグリッドサーチ・ランダムフォレストモデルにおける n_estimators   :  %d'  %RFC_grid.best_estimator_.n_estimators)
        
        print("グリッドサーチ・ランダムフォレスト")
        print(RFC_grid.best_estimator_)#最も性能がよかったランダムフォレストのインスタンス
        print('Best params: {}'.format(RFC_grid.best_params_)) 
        print('Best Score: {}'.format(RFC_grid.best_score_))#'criterion': 'entropy'の場合は小さいほうが良い、gridの場合は大きいほうがよい

        # 予測値算出
        #https://aiacademy.jp/media/?p=258
        pred_test = model2.predict(test_x)
        print( "\n [ テストデータ結果 ]" )
        print("適合率（Precision）, 再現率（Recall）, F値(f1-scoreのavg/totalの部分)")
        test_class = classification_report(test_y, pred_test,output_dict = True)
        print( test_class )

        print( "\n [ 混同行列 ]" )
        test_conf = confusion_matrix(test_y, pred_test)
        print( test_conf )

        print( "\n [ 正解率 ]" )#予測結果全体がどれくらい真の値と一致しているかを表す指標
        print( accuracy_score(test_y, pred_test) )
        accuracy_score_rdf = accuracy_score(test_y, pred_test)
        
        # feature = model2.feature_importances_
        # # 特徴量の名前ラベルを取得
        # label = df_list_train.columns[0:]
        # # 特徴量の重要度順（降順）に並べて表示
        # indices = np.argsort(feature)[::-1]
        # for i in range(len(feature)):
        #     print(str(i + 1) + "   " +
        #         str(label[indices[i]]) + "   " + str(feature[indices[i]]))    


        report_df = pd.DataFrame(train_class).T
        report_df.to_csv("output/" + 'rdf_train'+str(l)+'.csv')
        report_df = pd.DataFrame(test_class).T
        report_df.to_csv("output/" + 'rdf_test'+str(l)+'.csv')
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]

        param_list = []
        param_list.append(["rdf", str(l)])
        for key in best_model_params:
            param_list.append([key, best_model_params[key]])

        with open('parameter/' + 'parameter.txt', 'w') as o:
            for row in param_list:
                print(*row, sep=",", file=o)

        for j in range(len(train_conf)):
            for k in range(len(train_conf[1])):
                ws.cell(row = j+1,column = k+1,value = train_conf[j][k])
                ws.cell(row = j+7,column = k+1,value = test_conf[j][k])
        
        wb.save("output/" + 'rdf'+ str(l) + title + '.xlsx')
    return best_model_params, test_conf, accuracy_score_rdf


def SVMclass(
        train_x=None,
        train_y=None,
        test_x=None,
        test_y=None,
        title="hoge", 
        C=1, 
        kernel="rbf", 
        gamma="scale"
        ):
    #クラス間でのデータ数の不均衡を考慮
    #https://qiita.com/kento1109/items/1fc7488163b0f350f2fa
    poms=["workload"]
    for i in poms:
        tuned_parameters = [
            {'C': C, 'kernel': kernel, 'gamma': gamma}
        ]     
        #最適化の実行
        clf = GridSearchCV(
            SVC(class_weight='balanced'), # 識別器
            tuned_parameters, # 最適化したいパラメータセット 
            cv=5, # 交差検定の回数
            scoring='accuracy' ) # モデルの評価関数の指定
        print("clf",clf)
        clf.fit(train_x, train_y)

        #最適なモデルの選択-------------------------------------------------------------------------------------------------------------------------------
        best_model = clf.best_estimator_
        best_model_params=clf.best_params_
        print("最適なモデルのパラメータ",best_model_params)
        
        C_ = best_model_params["C"]
        kernel_ = best_model_params["kernel"]
        C.append(best_model_params["C"])
        kernel.append(best_model_params["kernel"])
        if kernel_=="rbf":
            gamma_=best_model_params["gamma"]
            gamma.append(best_model_params["gamma"])
            #モデルの構築
            model2=SVC(C=C_,kernel=kernel_,gamma=gamma_,class_weight='balanced',random_state=0)
        
        else:
            model2 = SVC(C=C_,kernel=kernel_,class_weight='balanced',random_state=0)
        print("model2",model2)
        
        # モデルのコンパイル
        #model_.compile(loss='mse',
        #              optimizer=tf.keras.optimizers.Adam(learning_rate=lr_),
        #              metrics=['mae', 'mse'])

        # モデルの学習
        history_ = model2.fit(train_x, train_y)
        print("モデルの学習",history_)

        with open("output/" + 'svc'+ str(l) +'.pickle', mode='wb') as f:  # with構文でファイルパスとバイナリ書き込みモードを設定
            pickle.dump(history_, f)   
        #--------------------------------------------------------------------------------------------------------------------------------------------------
        
        #各試行でのスコアを確認
        #print(clf.best_estimator_)#最も性能がよかったランダムフォレストのインスタンス
        #print('Best params: {}'.format(clf.best_params_)) 
        #print('Best Score: {}'.format(clf.best_score_))#'criterion': 'entropy'の場合は小さいほうが良い、gridの場合は大きいほうがよい
        
        # 予測値算出
        #https://aiacademy.jp/media/?p=258
        pred_train = model2.predict(train_x)
        print("グリッドサーチ・SVC")
        print( "\n [ 訓練データ結果 ]" )
        print("適合率（Precision）, 再現率（Recall）, F値(f1-scoreのavg/totalの部分)")
        train_class = classification_report(train_y, pred_train,output_dict = True)
        print( train_class )


        print( "\n [ 混同行列 ]" )
        train_conf = confusion_matrix(train_y, pred_train)
        print( train_conf )

        print( "\n [ 正解率 ]" )#予測結果全体がどれくらい真の値と一致しているかを表す指標
        print( accuracy_score(train_y, pred_train) )
        
        
        # 予測値算出
        #https://aiacademy.jp/media/?p=258
        pred_test= model2.predict(test_x)
        print("グリッドサーチ・SVC")
        print( "\n [ テストデータ結果 ]" )
        print("適合率（Precision）, 再現率（Recall）, F値(f1-scoreのavg/totalの部分)")
        test_class = classification_report(test_y, pred_test,output_dict = True)
        print( test_class )

        print( "\n [ 混同行列 ]" )
        test_conf = confusion_matrix(test_y, pred_test)
        print( test_conf )

        print( "\n [ 正解率 ]" )#予測結果全体がどれくらい真の値と一致しているかを表す指標
        print( accuracy_score(test_y, pred_test) )
        accuracy_score_svm = accuracy_score(test_y, pred_test)

        
        report_df = pd.DataFrame(train_class).T
        report_df.to_csv("output/" + 'svm_train'+str(l)+'.csv')
        report_df = pd.DataFrame(test_class).T
        report_df.to_csv("output/" + 'svm_test'+str(l)+'.csv')
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]

        param_list = []
        param_list.append(["svm", str(l)])
        for key in best_model_params:
            param_list.append([key, best_model_params[key]])

        with open('parameter/' + 'parameter.txt', 'w') as o:
            for row in param_list:
                print(*row, sep=",", file=o)

        for j in range(len(train_conf)):
            for k in range(len(train_conf[1])):
                ws.cell(row = j+1,column = k+1,value = train_conf[j][k])
                ws.cell(row = j+7,column = k+1,value = test_conf[j][k])
        
        wb.save("output/" + 'svm'+ str(l) + title + '.xlsx')
    return best_model_params, test_conf, accuracy_score_svm

def datasetting(l):
    df_list_test = []
    df_list_train = []
    files = glob.glob('now_all/*')
    test = [files[1],files[5],files[7],files[9],files[11],files[14],files[17],files[19],files[22],files[24]]
    for file in files:
        if file in test:
            sub_df = pd.read_csv(file)
            sub_wl = sub_df.pop("workload")
            sub_wl = sub_wl[l:]
            sub_wl = sub_wl.reset_index(drop=True)
            sub_df = pd.concat([sub_df, sub_wl],axis = 1)
            sub_df = sub_df.dropna()
            df_list_test.append(sub_df)
        else :
            sub_df = pd.read_csv(file)
            sub_wl = sub_df.pop("workload")
            sub_wl = sub_wl[l:]
            sub_wl = sub_wl.reset_index(drop=True)
            sub_df = pd.concat([sub_df, sub_wl],axis = 1)
            sub_df = sub_df.dropna()
            df_list_train.append(sub_df)
    df_list_train = pd.concat(df_list_train)
    df_list_test = pd.concat(df_list_test)
    print(df_list_train)
    print(df_list_test)

    train_x = df_list_train.drop("workload",axis=1)
    train_y= df_list_train[["workload"]]
    test_x=df_list_test.drop("workload",axis=1)
    test_y= df_list_test[["workload"]]

    sm = SMOTE()

    print(train_x)
    print(test_x)

    train_x, train_y = sm.fit_resample(train_x, train_y)
    test_x, test_y = sm.fit_resample(test_x, test_y)
    

    print(train_x)
    print(test_x)

    train3 = []
    test3 = []

    train2 = pd.concat([train_x,train_y], axis=1)
    test2 = pd.concat([test_x, test_y], axis=1)
    print(train2)
    print(test2)

    for j in range(1,6):
        train4 = train2[train2['workload'] == j]
        test4 = test2[test2['workload'] == j]
        train4 = train4[::5]
        test4 = test4[::5]    
        train3.append(train4)
        test3.append(test4)
        print(len(train4))

    train3 = pd.concat(train3)
    test3 = pd.concat(test3)
    print("________________________")

    print(train3)
    print(test3)

    train_x2 = train3.drop("workload",axis=1)
    train_y2= train3[["workload"]]
    test_x2=test3.drop("workload",axis=1)
    test_y2= test3[["workload"]]       

    train_x = np.array(train_x2)
    test_x = np.array(test_x2)
    train_y = np.array(train_y2)
    test_y = np.array(test_y2)
    return train_x, test_x, train_y, test_y


def mailing(subject, body):
    smtp_host = 'smtp.gmail.com'
    smtp_port = 465
    username = 'adc1120rk@gmail.com'
    password = ''
    from_address = 'adc1120rk@gmail.com'
    to_address = 'adc1120rk@gmail.com'
    subject = subject
    body = body
    message = ("From: %s\r\nTo: %s\r\nSubject: %s \r\n\r\n %s" % (from_address, to_address, subject, body))
    smtp = smtplib.SMTP_SSL(smtp_host, smtp_port)
    smtp.login(username, password)
    result = smtp.sendmail(from_address, to_address, message)


if __name__ == '__main__' :
    l_list = [0, 300, 600, 900]
    #l_list = [0]
    for l in l_list:
        train_x, test_x, train_y, test_y = datasetting(l)
        title = str(l)
        sec = l/10
        best_param_nn, test_conf_nn, accuracy_score_nn = NNclass(
                        train_x=train_x,
                        train_y=train_y,
                        test_x=test_x,
                        test_y=test_y,
                        title = title,
                        hidden_layer_sizes = [5], 
                        solver = ["adam"], 
                        activation = ["relu"], 
                        max_iter = [30])
        # best_param_rdf = RDFclass(title=str(l))
        # best_param_svm = SVMclass(title=str(l))
        accuracy_score_nn_str = str(accuracy_score_nn)
        best_param_nn_str = ' '.join(str(s) for s in best_param_nn)
        test_conf_nn_str1 = ' '.join(str(s) for s in test_conf_nn[0])
        test_conf_nn_str2 = ' '.join(str(s) for s in test_conf_nn[1])
        test_conf_nn_str3 = ' '.join(str(s) for s in test_conf_nn[2])
        test_conf_nn_str4 = ' '.join(str(s) for s in test_conf_nn[3])
        test_conf_nn_str5 = ' '.join(str(s) for s in test_conf_nn[4])
        body = ("accuracy score is %s \r\n Here is the confusion matrix \r\n%s \r\n %s\r\n %s\r\n %s\r\n %s" % (accuracy_score_nn_str, test_conf_nn_str1, test_conf_nn_str2, test_conf_nn_str3, test_conf_nn_str4, test_conf_nn_str5))
        print(body)
        subject = "NNmodel" + str(sec) + "seconds"
        mailing(subject, body)

        best_param_nn, test_conf_nn, accuracy_score_nn = RDFclass(
                        train_x=train_x,
                        train_y=train_y,
                        test_x=test_x,
                        test_y=test_y,
                        title = title,
                        n_estimators=[100, 200, 300, 400],
                        max_depth=[10, 30, 50, None]
                        max_features=["sqrt", "log2"]
                        max_iter = [30])
        # best_param_rdf = RDFclass(title=str(l))
        # best_param_svm = SVMclass(title=str(l))
        accuracy_score_nn_str = str(accuracy_score_nn)
        best_param_nn_str = ' '.join(str(s) for s in best_param_nn)
        test_conf_nn_str1 = ' '.join(str(s) for s in test_conf_nn[0])
        test_conf_nn_str2 = ' '.join(str(s) for s in test_conf_nn[1])
        test_conf_nn_str3 = ' '.join(str(s) for s in test_conf_nn[2])
        test_conf_nn_str4 = ' '.join(str(s) for s in test_conf_nn[3])
        test_conf_nn_str5 = ' '.join(str(s) for s in test_conf_nn[4])
        body = ("accuracy score is %s \r\n Here is the confusion matrix \r\n%s \r\n %s\r\n %s\r\n %s\r\n %s" % (accuracy_score_nn_str, test_conf_nn_str1, test_conf_nn_str2, test_conf_nn_str3, test_conf_nn_str4, test_conf_nn_str5))
        print(body)
        subject = "rdfmodel" + str(sec) + "seconds"
        mailing(subject, body)

        best_param_nn, test_conf_nn, accuracy_score_nn = SVMclass(
                        train_x=train_x,
                        train_y=train_y,
                        test_x=test_x,
                        test_y=test_y,
                        title = title,
                        C = [0.1, 0.3, 1, 3, 10], 
                        gamma = [0.01, 0.03, 0.1, 0.3, 1, 3, 10], 
                        max_iter = [100])
        # best_param_rdf = RDFclass(title=str(l))
        # best_param_svm = SVMclass(title=str(l))
        accuracy_score_nn_str = str(accuracy_score_nn)
        best_param_nn_str = ' '.join(str(s) for s in best_param_nn)
        test_conf_nn_str1 = ' '.join(str(s) for s in test_conf_nn[0])
        test_conf_nn_str2 = ' '.join(str(s) for s in test_conf_nn[1])
        test_conf_nn_str3 = ' '.join(str(s) for s in test_conf_nn[2])
        test_conf_nn_str4 = ' '.join(str(s) for s in test_conf_nn[3])
        test_conf_nn_str5 = ' '.join(str(s) for s in test_conf_nn[4])
        body = ("accuracy score is %s \r\n Here is the confusion matrix \r\n%s \r\n %s\r\n %s\r\n %s\r\n %s" % (accuracy_score_nn_str, test_conf_nn_str1, test_conf_nn_str2, test_conf_nn_str3, test_conf_nn_str4, test_conf_nn_str5))
        print(body)
        subject = "svmmodel" + str(sec) + "seconds"
        mailing(subject, body)


